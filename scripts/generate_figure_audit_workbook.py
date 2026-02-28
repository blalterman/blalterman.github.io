#!/usr/bin/env python3
"""
Generate Excel audit workbook for figure corpus verification.

Creates one sheet per paper with columns for:
- thumbnail: Small image of the figure for visual verification
- corpus_figure_id: Figure ID from paper_metadata.json
- correct_figure_id: User enters corrected ID if wrong (manual entry)
- pdf_exists: Whether PDF exists in corpus
- svg_exists: Whether SVG exists in public/
- in_registry: Whether figure is in figure-registry.json
- multi_panel: Value from metadata
- panel_count: Value from metadata
- notes: User comments (manual entry)

Output: review-docs/figure-audit.xlsx
"""

import json
import sys
import tempfile
import io
from pathlib import Path

# Add parent dir to path for utils import
sys.path.insert(0, str(Path(__file__).parent))
from utils import get_repo_root, get_public_data_dir

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("Error: openpyxl is required. Install with: conda install openpyxl")
    sys.exit(1)

try:
    from pdf2image import convert_from_path
    from PIL import Image
    PDF_SUPPORT = True
except ImportError:
    print("Warning: pdf2image not available. Thumbnails will be skipped.")
    PDF_SUPPORT = False

# Thumbnail settings
THUMB_WIDTH = 150  # pixels
THUMB_HEIGHT = 120  # pixels
ROW_HEIGHT = 95  # Excel row height (points)


def create_thumbnail(pdf_path: Path, temp_dir: Path, paper_id: str) -> Path | None:
    """Convert first page of PDF to thumbnail image."""
    if not PDF_SUPPORT or not pdf_path.exists():
        return None

    try:
        # Convert first page only, at low DPI for speed
        images = convert_from_path(
            pdf_path,
            first_page=1,
            last_page=1,
            dpi=72,
            size=(THUMB_WIDTH * 2, None)  # 2x for better quality when scaled
        )

        if not images:
            return None

        img = images[0]

        # Create thumbnail
        img.thumbnail((THUMB_WIDTH, THUMB_HEIGHT), Image.Resampling.LANCZOS)

        # Save to temp file - include paper_id to avoid collisions
        thumb_path = temp_dir / f"{paper_id}_{pdf_path.stem}_thumb.png"
        img.save(thumb_path, "PNG")

        return thumb_path

    except Exception as e:
        print(f"    Warning: Could not create thumbnail for {pdf_path.name}: {e}")
        return None


def load_corpus_data(repo_root: Path) -> dict:
    """Load paper metadata from research-corpus."""
    corpus_dir = repo_root / "research-corpus" / "papers"
    papers = {}

    for paper_dir in sorted(corpus_dir.iterdir()):
        if not paper_dir.is_dir() or paper_dir.name.startswith('.'):
            continue

        metadata_file = paper_dir / "paper_metadata.json"
        if not metadata_file.exists():
            continue

        with open(metadata_file, 'r') as f:
            data = json.load(f)

        paper_id = paper_dir.name
        figures_dir = paper_dir / "figures"

        # Get list of actual PDF files
        pdf_files = []
        if figures_dir.exists():
            pdf_files = sorted([f.name for f in figures_dir.glob("fig_*.pdf")])

        papers[paper_id] = {
            "metadata": data,
            "pdf_files": pdf_files,
            "figures_dir": figures_dir
        }

    return papers


def load_figure_registry(repo_root: Path) -> dict:
    """Load figure-registry.json and index by figure ID."""
    registry_file = get_public_data_dir() / "figure-registry.json"

    if not registry_file.exists():
        return {}

    with open(registry_file, 'r') as f:
        registry = json.load(f)

    # Registry is already indexed by "paper_id/figure_id"
    return registry


def check_svg_exists(repo_root: Path, paper_id: str, figure_id: str) -> bool:
    """Check if SVG exists in public/papers/{paper_id}/figures/."""
    svg_path = repo_root / "public" / "papers" / paper_id / "figures" / f"{figure_id}.svg"
    return svg_path.exists()


def create_audit_workbook(papers: dict, registry: dict, repo_root: Path, temp_dir: Path) -> Workbook:
    """Create Excel workbook with one sheet per paper."""
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    issue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers with thumbnail column first
    headers = [
        ("A", "thumbnail", 22),  # ~150 pixels
        ("B", "corpus_figure_id", 18),
        ("C", "correct_figure_id", 18),
        ("D", "pdf_exists", 10),
        ("E", "svg_exists", 10),
        ("F", "in_registry", 10),
        ("G", "multi_panel", 10),
        ("H", "panel_count", 10),
        ("I", "technical_caption", 50),
        ("J", "notes", 30)
    ]

    for paper_id, paper_data in sorted(papers.items()):
        print(f"  Processing {paper_id}...")

        # Create sheet (truncate name if needed, Excel limit is 31 chars)
        sheet_name = paper_id[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Add headers
        for col, header, width in headers:
            cell = ws[f"{col}1"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[col].width = width

        # Get metadata figures
        metadata_figures = {
            fig["figure_id"]: fig
            for fig in paper_data["metadata"].get("figures", [])
        }

        # Combine metadata figures and PDF files to catch mismatches
        all_figure_ids = set(metadata_figures.keys())
        for pdf_file in paper_data["pdf_files"]:
            fig_id = pdf_file.replace(".pdf", "")
            all_figure_ids.add(fig_id)

        def sort_key(fig_id: str) -> tuple:
            """Sort fig_1, fig_2, ... fig_10, then non-standard names."""
            import re
            match = re.match(r'fig_(\d+)$', fig_id)
            if match:
                return (0, int(match.group(1)), "")
            # Non-standard names sort after standard ones
            return (1, 0, fig_id)

        row = 2
        for fig_id in sorted(all_figure_ids, key=sort_key):
            fig_meta = metadata_figures.get(fig_id, {})
            registry_id = f"{paper_id}/{fig_id}"

            pdf_exists = f"{fig_id}.pdf" in paper_data["pdf_files"]
            svg_exists = check_svg_exists(repo_root, paper_id, fig_id)
            in_registry = registry_id in registry

            # Set row height to accommodate thumbnail
            ws.row_dimensions[row].height = ROW_HEIGHT

            # Try to add thumbnail
            if pdf_exists and PDF_SUPPORT:
                pdf_path = paper_data["figures_dir"] / f"{fig_id}.pdf"
                thumb_path = create_thumbnail(pdf_path, temp_dir, paper_id)
                if thumb_path:
                    try:
                        img = XLImage(str(thumb_path))
                        img.width = THUMB_WIDTH
                        img.height = THUMB_HEIGHT
                        # Anchor to cell
                        ws.add_image(img, f"A{row}")
                    except Exception as e:
                        print(f"    Warning: Could not add image for {fig_id}: {e}")

            # Write data
            ws[f"B{row}"] = fig_id
            ws[f"C{row}"] = ""  # User fills in if correction needed
            ws[f"D{row}"] = "YES" if pdf_exists else "NO"
            ws[f"E{row}"] = "YES" if svg_exists else "NO"
            ws[f"F{row}"] = "YES" if in_registry else "NO"
            ws[f"G{row}"] = "YES" if fig_meta.get("multi_panel") else "NO" if fig_meta.get("multi_panel") is not None else ""
            ws[f"H{row}"] = fig_meta.get("panel_count", "")
            caption = fig_meta.get("technical_caption", "") or ""
            ws[f"I{row}"] = caption[:300] + "..." if len(caption) > 300 else caption
            ws[f"J{row}"] = ""  # User notes

            # Apply conditional formatting
            for col in ["D", "E", "F"]:
                cell = ws[f"{col}{row}"]
                if cell.value == "NO":
                    cell.fill = issue_fill
                else:
                    cell.fill = ok_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply borders and alignment to other cells
            for col in ["A", "B", "C", "G", "H", "I", "J"]:
                ws[f"{col}{row}"].border = thin_border
                ws[f"{col}{row}"].alignment = Alignment(vertical="center", wrap_text=(col == "I"))

            row += 1

        # Freeze header row and thumbnail column
        ws.freeze_panes = "B2"

    # Remove default sheet if it still exists
    if default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    return wb


def main():
    """Generate the figure audit workbook."""
    repo_root = get_repo_root()

    print("Loading corpus data...")
    papers = load_corpus_data(repo_root)
    print(f"  Found {len(papers)} papers")

    print("Loading figure registry...")
    registry = load_figure_registry(repo_root)
    print(f"  Found {len(registry)} registered figures")

    print("Creating audit workbook with thumbnails...")

    # Create temp directory for thumbnails
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        wb = create_audit_workbook(papers, registry, repo_root, temp_path)

        output_path = repo_root / "review-docs" / "figure-audit.xlsx"
        output_path.parent.mkdir(exist_ok=True)
        wb.save(output_path)

    print(f"\n{output_path}")
    print("\nWorkbook created with sheets:")
    for paper_id in sorted(papers.keys()):
        fig_count = len(papers[paper_id]["pdf_files"])
        meta_count = len(papers[paper_id]["metadata"].get("figures", []))
        mismatch = " (MISMATCH)" if fig_count != meta_count else ""
        print(f"  - {paper_id}: {fig_count} PDFs, {meta_count} in metadata{mismatch}")

    print("\nNext steps:")
    print("  1. Open figure-audit.xlsx in Excel")
    print("  2. Review each sheet - thumbnails show actual figure content")
    print("  3. For incorrect figure IDs, enter correct value in 'correct_figure_id' column")
    print("     (Use format like: 1a, 2b, 3, DELETE, etc.)")
    print("  4. Add notes for any issues found")
    print("  5. Save and return to Claude for processing")


if __name__ == "__main__":
    main()
